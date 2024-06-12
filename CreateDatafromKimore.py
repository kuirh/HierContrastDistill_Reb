import os
import csv
import openpyxl
import numpy as np



def check_directory(directory_path):
    raw_directory = os.path.join(directory_path, "Raw")
    label_directory = os.path.join(directory_path, "Label")

    has_depth_file = False
    has_joint_position_file = False
    has_joint_orientation_file = False
    has_clinical_assessment_file = False

    # # 检查Raw子目录中是否存在以depth开头的文件
    raw_files = os.listdir(raw_directory)
    # for file in raw_files:
    #     if file.startswith("depth"):
    #         has_depth_file = True
    #         break

    # 检查Raw子目录中是否存在以JointPosition和JointOrientation开头的文件
    for file in raw_files:
        if file.startswith("JointPosition"):
            has_joint_position_file = True
        if file.startswith("JointOrientation"):
            has_joint_orientation_file = True

    # 检查Label子目录中是否存在以ClinicalAssessment开头的文件
    label_files = os.listdir(label_directory)
    for file in label_files:
        if file.startswith("ClinicalAssessment"):
            has_clinical_assessment_file = True
            break

    if  has_joint_position_file and has_joint_orientation_file and has_clinical_assessment_file:
        return True
    else:
        return False




path = "../KiMoRe"
kinect_joints = ["spinebase", "spinemid", "neck", "head",
                 "shoulderleft", "elbowleft", "wristleft",
                 "handleft", "shoulderright", "elbowright",
                 "wristright", "handright", "hipleft", "kneeleft",
                 "ankleleft", "footleft", "hipright", "kneeright",
                 "ankleright", "footright", "spineshoulder", "handtipleft",
                 "thumbleft", "handtipright", "thumbright"]

enable_kinect_joints = True
enable_slice_list = False

train_data = {
    'ex1': {'train_x_ori': [], 'train_x_pos': [], 'train_y': []},
    'ex2': {'train_x_ori': [], 'train_x_pos': [], 'train_y': []},
    'ex3': {'train_x_ori': [], 'train_x_pos': [], 'train_y': []},
    'ex4': {'train_x_ori': [], 'train_x_pos': [], 'train_y': []},
    'ex5': {'train_x_ori': [], 'train_x_pos': [], 'train_y': []},
}


def fill_train_data(exercise_num, data_ori, data_pos,y):
    key = f'ex{exercise_num}'  # 根据练习编号生成键名
    if key in train_data:  # 检查键名是否存在
        train_data[key]['train_x_ori'].append(data_ori)
        train_data[key]['train_x_pos'].append(data_pos)
        train_data[key]['train_y'].append(y)



for (root, dirs, files) in os.walk(path):  #

    # if current directory contains "Raw", extract data
    if "Raw" in dirs and check_directory(root):

        new_dict = {}
        pos={}
        ori={}
        # get exercise number
        new_dict["Exercise"] = int(root[-1])

        # extract raw data
        raw_files = os.listdir(os.path.join(root, "Raw"))

        for file in raw_files:

            file_path = os.path.join(os.path.join(root, "Raw"), file)
            csv_file = open(file_path, newline='')
            csv_reader = csv.reader(csv_file)

            if file.startswith("JointOrientation"):

                if enable_kinect_joints:
                    for joint in kinect_joints:
                        ori[joint + "-o"] = []
                        new_dict['ori'] = ori

                    for row in csv_reader:
                        for i in range(len(kinect_joints)):
                            if len(row) > 0:
                                ori[kinect_joints[i] + "-o"].append(np.array(row[(4 * i):(4 * i + 4)]))
                else:
                    new_dict["JointOrientation"] = []
                    for row in csv_reader:
                        if len(new_dict["JointOrientation"]) >= 182 and enable_slice_list:
                            break
                        elif len(row) > 0:
                            if '' in row:
                                row.remove('')
                            new_dict["JointOrientation"].append(np.array([float(i) for i in row]))
                    np.array(new_dict["JointOrientation"])


            elif file.startswith("JointPosition"):

                if enable_kinect_joints:
                    for joint in kinect_joints:
                        pos[joint + "-p"] = []

                    for row in csv_reader:
                        for i in range(len(kinect_joints)):
                            if len(row) > 0:
                                pos[kinect_joints[i] + "-p"].append(np.array(row[(4 * i):(4 * i + 3)]))
                                new_dict['pos']=pos
                else:
                    new_dict["JointPosition"] = []
                    for row in csv_reader:
                        if len(new_dict["JointPosition"]) >= 182 and enable_slice_list:
                            break
                        elif len(row) > 0:
                            if '' in row:
                                row.remove('')
                            new_dict["JointPosition"].append(np.array([float(i) for i in row]))
                    np.array(new_dict["JointPosition"])

            elif file.startswith("TimeStamp"):

                new_dict["Timestamps"] = []
                for row in csv_reader:
                    if len(new_dict["Timestamps"]) >= 182 and enable_slice_list:
                        break
                    elif len(row) > 0:
                        if '' in row:
                            row.remove('')
                        new_dict["Timestamps"].append(row)





        # extract data labels
        label_files = os.listdir(os.path.join(root, "Label"))
        for file in label_files:

            file_path = os.path.join(os.path.join(root, "Label"), file)
            book = openpyxl.load_workbook(file_path)
            sheet = book.active

            if file.startswith("SuppInfo"):
                for i in range(1, sheet.max_column):
                    t = sheet.cell(1, i).value
                    v = sheet.cell(2, i).value
                    new_dict[t] = v

            elif file.startswith("ClinicalAssessment"):
                new_dict["cTS"] = sheet.cell(2, new_dict["Exercise"] + 1).value
                new_dict["cPO"] = sheet.cell(2, new_dict["Exercise"] + 6).value
                new_dict["cCF"] = sheet.cell(2, new_dict["Exercise"] + 11).value
        # append exercise to data
        n = len(next(iter(ori.values())))
        oriarry = []
        pndarry = []
        for l in sorted(ori.keys()):
            oriarry.append(np.stack(ori[l], axis=0))

        oriarry = np.stack(oriarry, axis=1)
        #oriarry= Timecorrect.timecorrect(oriarry)
        for l in sorted(pos.keys()):
            pndarry.append(np.stack(pos[l], axis=0))
        pndarry = np.stack(pndarry, axis=1)
        #pndarry = Timecorrect.timecorrect(pndarry)


        if new_dict.get("cTS") and len(pndarry)==len(oriarry):
            num_sections = len(pndarry) // 100

            new_len = (num_sections) * 100

            arr1_truncated = pndarry[:new_len].reshape(-1, 100, 25, 3)
            arr2_truncated = oriarry[:new_len].reshape(-1, 100, 25, 4)
            label_truncated = np.full((num_sections, 1), new_dict["cTS"])
            fill_train_data(new_dict["Exercise"], arr2_truncated, arr1_truncated, label_truncated)

#print(train_data)

for key, value in train_data.items():
    main_dir_path = os.path.join('Dataset_cTS', str(key))
    os.makedirs(main_dir_path, exist_ok=True)
    if value['train_x_ori'] and value['train_x_pos'] and value['train_y']:
        a=np.concatenate(value['train_x_ori'], axis=0)
        b=np.concatenate(value['train_x_pos'], axis=0)
        c=np.concatenate(value['train_y'], axis=0)
        np.save(os.path.join(main_dir_path, 'train_x_ori.npy'), a)
        np.save(os.path.join(main_dir_path, 'train_x_pos.npy'), b)
        np.save(os.path.join(main_dir_path, 'train_y.npy'), c)

        # 每个部分的label都是原来的整数值

#
#
#
#
#
#
# list_JointPosition = []
# list_JointOrientation = []
# list_cCF = []
# list_depth= []
# # 填充数组
# for i, d in enumerate(data):
#     list_JointPosition.append( d['JointPosition'])
#     list_JointOrientation.append(d['JointOrientation'])
#     list_cCF.append(d['cCF'])
#     list_depth.append(d['depth'])
    #merged_arr = np.concatenate(list_depth, axis=0)


# with open('JointPosition.pkl', 'wb') as f:
#     pickle.dump(list_JointPosition, f)
#
# with open('JointOrientation.pkl', 'wb') as f:
#     pickle.dump(list_JointOrientation, f)
#
# with open('cCF.pkl', 'wb') as f:
#     pickle.dump(list_cCF, f)
#
# np.savetxt('array.txt', merged_arr)
